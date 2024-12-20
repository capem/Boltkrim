import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import json
import os
import fitz  # PyMuPDF
import pandas as pd
from fuzzywuzzy import process
from fuzzywuzzy import fuzz
from pathlib import Path
import openpyxl
from PIL import Image, ImageTk
import io
import shutil
import tempfile

class FuzzySearchFrame(ttk.Frame):
    """A frame that provides fuzzy search functionality with a text entry and listbox.
    
    This widget allows users to search through a list of values using fuzzy matching,
    displaying the best matches in a scrollable listbox.
    """
    
    def __init__(
        self,
        master: tk.Widget,
        values: list[str] | None = None,
        search_threshold: int = 65,
        identifier: str | None = None,
        **kwargs
    ) -> None:
        """Initialize the FuzzySearchFrame.
        
        Args:
            master: The parent widget
            values: Initial list of values to search through
            search_threshold: Minimum score (0-100) for fuzzy matches
            identifier: Unique identifier for debugging purposes
            **kwargs: Additional keyword arguments for the Frame
        """
        super().__init__(master, **kwargs)
        
        self.all_values = [str(v) for v in (values or []) if v is not None]
        self.search_threshold = max(0, min(100, search_threshold))  # Clamp between 0 and 100
        self.identifier = identifier or 'unnamed'
        
        # Debouncing variables
        self._prev_value = ''
        self._after_id = None
        self._ignore_next_keyrelease = False
        self._debounce_delay = 0  # Changed from 150 to 0 to remove delay
        self._focus_after_id = None
        
        self._create_widgets()
        self._bind_events()
        self._update_listbox()
        
        # Schedule focus set after widget is fully realized
        self.after(100, self._ensure_focus)
        
    def _create_widgets(self) -> None:
        """Create and configure all child widgets."""
        # Entry widget
        self.entry = ttk.Entry(self)
        self.entry.pack(fill='x', padx=2, pady=2)
        
        # Bind focus-related events
        self.entry.bind('<FocusIn>', self._on_focus_in)
        self.entry.bind('<FocusOut>', self._on_focus_out)
        
        # Listbox frame with scrollbar
        listbox_frame = ttk.Frame(self)
        listbox_frame.pack(fill='both', expand=True, padx=2)
        
        # Listbox
        self.listbox = tk.Listbox(
            listbox_frame,
            height=5,
            exportselection=False,
            selectmode=tk.SINGLE
        )
        self.listbox.pack(side='left', fill='both', expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(
            listbox_frame,
            orient='vertical',
            command=self.listbox.yview
        )
        scrollbar.pack(side='right', fill='y')
        self.listbox.configure(yscrollcommand=scrollbar.set)
        
    def _bind_events(self) -> None:
        """Bind widget events to their handlers."""
        self.entry.bind('<KeyRelease>', self._on_keyrelease)
        self.listbox.bind('<<ListboxSelect>>', self._on_select)
        self.listbox.bind('<Button-1>', lambda e: self.after(50, self._on_select, e))  # Add single-click binding with slight delay
        self.listbox.bind('<Double-Button-1>', self._on_select)  # Add double-click binding
        self.entry.bind('<Return>', lambda e: self._select_top_match())
        self.entry.bind('<Down>', lambda e: self._focus_listbox())
        self.entry.bind('<Tab>', self._handle_tab)  # Add Tab key binding
        
    def set_values(self, values: list[str] | None) -> None:
        """Update the list of searchable values.
        
        Args:
            values: New list of values to search through
        """
        self.all_values = [str(v) for v in (values or []) if v is not None]
        self.entry.delete(0, tk.END)
        self._update_listbox()
        
    def get(self) -> str:
        """Get the current entry text.
        
        Returns:
            The current text in the entry widget
        """
        return self.entry.get()
        
    def set(self, value: str) -> None:
        """Set the entry text.
        
        Args:
            value: Text to set in the entry widget
        """
        self.entry.delete(0, tk.END)
        self.entry.insert(0, str(value))
        
    def _on_keyrelease(self, event: tk.Event) -> None:
        """Handle key release events in the entry widget."""
        if self._ignore_next_keyrelease:
            self._ignore_next_keyrelease = False
            return
            
        # Update immediately without debouncing
        self._update_listbox()
        
    def _update_listbox(self) -> None:
        """Update the listbox with fuzzy search results."""
        current_value = self.entry.get().strip()
        
        # Clear current listbox
        self.listbox.delete(0, tk.END)
        
        # If empty, show all values
        if not current_value:
            for value in self.all_values:
                self.listbox.insert(tk.END, value)
            return
            
        try:
            # Convert search term to lowercase for case-insensitive matching
            search_lower = current_value.lower()
            
            # 1. Exact matches (case-insensitive)
            exact_matches = [v for v in self.all_values if v.lower() == search_lower]
            
            # 2. Prefix matches (prioritized by length)
            prefix_matches = [
                v for v in self.all_values 
                if v.lower().startswith(search_lower) 
                and v not in exact_matches
            ]
            # Sort prefix matches by length (shorter first) then alphabetically
            prefix_matches.sort(key=lambda x: (len(x), x.lower()))
            
            # 3. Contains matches (words that contain the search term)
            contains_matches = [
                v for v in self.all_values 
                if search_lower in v.lower() 
                and v not in exact_matches 
                and v not in prefix_matches
            ]
            
            # 4. Fuzzy matches for remaining items
            remaining_values = [
                v for v in self.all_values 
                if v not in exact_matches 
                and v not in prefix_matches 
                and v not in contains_matches
            ]
            
            # Use multiple fuzzy matching algorithms for better results
            token_matches = process.extract(
                current_value,
                remaining_values,
                limit=10,
                scorer=fuzz.token_sort_ratio
            )
            
            partial_matches = process.extract(
                current_value,
                remaining_values,
                limit=10,
                scorer=fuzz.partial_ratio
            )
            
            # Combine and deduplicate fuzzy matches
            fuzzy_matches = {}
            for match, score in token_matches + partial_matches:
                if match not in fuzzy_matches or score > fuzzy_matches[match]:
                    fuzzy_matches[match] = score
            
            # Filter and sort fuzzy matches
            filtered_fuzzy = [
                (match, score) 
                for match, score in fuzzy_matches.items() 
                if score >= self.search_threshold
            ]
            filtered_fuzzy.sort(key=lambda x: (-x[1], x[0].lower()))  # Sort by score (desc) then alphabetically
            
            # Add matches to listbox in priority order
            # 1. Exact matches
            for match in exact_matches:
                self.listbox.insert(tk.END, match)
                
            # 2. Prefix matches
            for match in prefix_matches:
                self.listbox.insert(tk.END, match)
                
            # 3. Contains matches
            for match in contains_matches:
                self.listbox.insert(tk.END, match)
                
            # 4. Fuzzy matches
            for match, _ in filtered_fuzzy:
                if match not in exact_matches and match not in prefix_matches and match not in contains_matches:
                    self.listbox.insert(tk.END, match)
                    
        except Exception as e:
            print(f"Error in fuzzy search ({self.identifier}): {str(e)}")
            # Fall back to simple substring matching
            for value in self.all_values:
                if current_value.lower() in value.lower():
                    self.listbox.insert(tk.END, value)
                    
    def _on_select(self, event: tk.Event) -> None:
        """Handle selection events in the listbox."""
        if not self.listbox.size():  # If listbox is empty, do nothing
            return
            
        selection = self.listbox.curselection()
        if selection:
            value = self.listbox.get(selection[0])
            self._ignore_next_keyrelease = True
            self.set(value)
            # Generate a virtual event that can be bound by parent widgets
            self.event_generate('<<ValueSelected>>')
            self.entry.focus_set()
            
    def _select_top_match(self) -> None:
        """Select the top match in the listbox when Enter is pressed."""
        if self.listbox.size() > 0:
            value = self.listbox.get(0)
            self._ignore_next_keyrelease = True
            self.set(value)
            
    def _focus_listbox(self) -> None:
        """Move focus to the listbox when Down arrow is pressed."""
        if self.listbox.size() > 0:
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(0)
            self.listbox.focus_set()
            
    def _ensure_focus(self) -> None:
        """Ensure the entry widget has focus."""
        if not self.entry.focus_get():
            self.entry.focus_force()
            # Schedule another check
            self._focus_after_id = self.after(100, self._ensure_focus)
            
    def _on_focus_in(self, event=None) -> None:
        """Handle focus-in event."""
        if self._focus_after_id:
            self.after_cancel(self._focus_after_id)
            self._focus_after_id = None
            
    def _on_focus_out(self, event=None) -> None:
        """Handle focus-out event."""
        # If we lose focus, try to get it back after a short delay
        # This helps with the initial focus issues
        self._focus_after_id = self.after(100, self._ensure_focus)
        
    def _handle_tab(self, event=None) -> None:
        """Handle Tab key press: select first result and move to next widget."""
        if self.listbox.size() > 0:
            # Select the first item
            value = self.listbox.get(0)
            self._ignore_next_keyrelease = True
            self.set(value)
            # Generate the ValueSelected event
            self.event_generate('<<ValueSelected>>')
            
        # Prevent default Tab behavior
        if event:
            event.widget.tk_focusNext().focus()
            return "break"
        
class ErrorDialog(tk.Toplevel):
    def __init__(self, parent, title, message):
        super().__init__(parent)
        self.title(title)
        
        # Make dialog modal
        self.transient(parent)
        self.grab_set()
        
        # Calculate position
        window_width = 500
        window_height = 300
        screen_width = parent.winfo_screenwidth()
        screen_height = parent.winfo_screenheight()
        center_x = int(screen_width/2 - window_width/2)
        center_y = int(screen_height/2 - window_height/2)
        
        # Set size and position
        self.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        self.minsize(window_width, window_height)
        
        # Add message
        message_frame = ttk.Frame(self)
        message_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Add scrolled text widget for error message
        self.text_widget = scrolledtext.ScrolledText(
            message_frame, 
            wrap=tk.WORD, 
            width=50, 
            height=10
        )
        self.text_widget.pack(fill='both', expand=True, padx=5, pady=5)
        self.text_widget.insert('1.0', message)
        self.text_widget.configure(state='disabled')
        
        # Add buttons
        button_frame = ttk.Frame(self)
        button_frame.pack(fill='x', padx=10, pady=(0, 10))
        
        copy_button = ttk.Button(
            button_frame, 
            text="Copy to Clipboard",
            command=self.copy_to_clipboard
        )
        copy_button.pack(side='left', padx=5)
        
        close_button = ttk.Button(
            button_frame, 
            text="Close",
            command=self.destroy
        )
        close_button.pack(side='right', padx=5)
        
        # Bind escape key to close
        self.bind('<Escape>', lambda e: self.destroy())
        
        # Center the dialog
        self.update_idletasks()
        
        # Make dialog resizable
        self.resizable(True, True)
        
    def copy_to_clipboard(self):
        error_text = self.text_widget.get('1.0', tk.END)
        self.clipboard_clear()
        self.clipboard_append(error_text)
        
class FileOrganizerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("File Organizer")
        self.root.geometry("1200x800")
        
        # Default configuration
        self.default_config = {
            'source_folder': '',
            'processed_folder': '',
            'excel_file': '',
            'excel_sheet': '',
            'filter1_column': '',  
            'filter2_column': ''   
        }
        
        # Initialize configuration with defaults
        self.config = self.default_config.copy()
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Create tabs
        self.processing_tab = ttk.Frame(self.notebook)
        self.config_tab = ttk.Frame(self.notebook)
        
        self.notebook.add(self.processing_tab, text='Processing')
        self.notebook.add(self.config_tab, text='Configuration')
        
        # Create status bar
        self.status_bar = ttk.Label(root, text="", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Initialize zoom level
        self.zoom_level = 1
        
        # Initialize variables
        self.current_pdf = None
        self.excel_data = None
        self.all_values1 = []
        self.all_values2 = []
        
        # Initialize UI components for processing tab
        self.processing_filter1_frame = None  
        self.processing_filter2_frame = None  
        
        # Initialize UI components for config tab
        self.config_filter1_frame = None  
        self.config_filter2_frame = None  
        
        # Initialize tabs
        self.setup_config_tab()
        self.setup_processing_tab()
        
        # Load saved configuration
        self.load_config()
        
        # Load initial data if configuration exists
        if self.config['excel_file'] and self.config['excel_sheet']:
            self.load_excel_data()
            
        # Load initial PDF if source folder exists
        if self.config['source_folder'] and os.path.exists(self.config['source_folder']):
            self.load_next_pdf()
        
        # Bind keyboard shortcuts
        self.root.bind('<Control-s>', lambda e: self.save_config())
        self.root.bind('<Control-n>', lambda e: self.load_next_pdf())
        self.root.bind('<Return>', lambda e: self.process_current_file())
        self.root.bind('<Control-plus>', lambda e: self.zoom_in())
        self.root.bind('<Control-minus>', lambda e: self.zoom_out())
        
    def setup_config_tab(self):
        # Create and pack widgets with proper spacing
        ttk.Label(self.config_tab, text="Configuration", font=('Helvetica', 14, 'bold')).pack(pady=10)
        
        # Source Folder
        frame = ttk.Frame(self.config_tab)
        frame.pack(fill='x', padx=20, pady=5)
        ttk.Label(frame, text="Source Folder:").pack(side='left')
        self.source_folder_entry = ttk.Entry(frame, width=50)
        self.source_folder_entry.pack(side='left', padx=5)
        self.source_folder_entry.insert(0, self.config['source_folder'])
        ttk.Button(frame, text="Browse", command=self.select_source_folder).pack(side='left')
        
        # Processed Folder
        frame = ttk.Frame(self.config_tab)
        frame.pack(fill='x', padx=20, pady=5)
        ttk.Label(frame, text="Processed Folder:").pack(side='left')
        self.processed_folder_entry = ttk.Entry(frame, width=50)
        self.processed_folder_entry.pack(side='left', padx=5)
        self.processed_folder_entry.insert(0, self.config['processed_folder'])
        ttk.Button(frame, text="Browse", command=self.select_processed_folder).pack(side='left')
        
        # Excel File
        frame = ttk.Frame(self.config_tab)
        frame.pack(fill='x', padx=20, pady=5)
        ttk.Label(frame, text="Excel File:").pack(side='left')
        self.excel_file_entry = ttk.Entry(frame, width=50)
        self.excel_file_entry.pack(side='left', padx=5)
        self.excel_file_entry.insert(0, self.config['excel_file'])
        ttk.Button(frame, text="Browse", command=self.select_excel_file).pack(side='left')
        
        # Excel Sheet
        frame = ttk.Frame(self.config_tab)
        frame.pack(fill='x', padx=20, pady=5)
        ttk.Label(frame, text="Excel Sheet:").pack(side='left')
        self.sheet_combobox = ttk.Combobox(frame, width=47)
        self.sheet_combobox.pack(side='left', padx=5)
        self.sheet_combobox.bind('<<ComboboxSelected>>', lambda e: self.update_column_lists())
        
        # Column Configurations
        frame = ttk.Frame(self.config_tab)
        frame.pack(fill='x', padx=20, pady=5)
        ttk.Label(frame, text="First Column:").pack(side='left')
        self.config_filter1_frame = FuzzySearchFrame(frame, width=47, identifier='config_filter1')
        self.config_filter1_frame.pack(side='left', padx=5, fill='x', expand=True)
        
        frame = ttk.Frame(self.config_tab)
        frame.pack(fill='x', padx=20, pady=5)
        ttk.Label(frame, text="Second Column:").pack(side='left')
        self.config_filter2_frame = FuzzySearchFrame(frame, width=47, identifier='config_filter2')
        self.config_filter2_frame.pack(side='left', padx=5, fill='x', expand=True)
        
        # Save Button
        ttk.Button(self.config_tab, text="Save Configuration", 
                  command=self.save_config).pack(pady=20)
        
        # Initialize Excel data if file exists
        if self.config['excel_file']:
            self.update_sheet_list()
            # Update column lists after a short delay to ensure widgets are ready
            self.root.after(100, self.update_column_lists)
            
    def setup_processing_tab(self):
        # Create main container
        container = ttk.Frame(self.processing_tab)
        container.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Create left frame for PDF viewer with scrollbars
        self.pdf_frame = ttk.Frame(container)
        self.pdf_frame.pack(side='left', fill='both', expand=True)
        
        # Create right frame for controls
        controls_frame = ttk.Frame(container)
        controls_frame.pack(side='right', fill='y', padx=10)
        
        # Add zoom controls
        zoom_frame = ttk.Frame(controls_frame)
        zoom_frame.pack(fill='x', pady=5)
        ttk.Button(zoom_frame, text="-", width=3,
                  command=self.zoom_out).pack(side='left', padx=2)
        self.zoom_label = ttk.Label(zoom_frame, text="100%")
        self.zoom_label.pack(side='left', padx=5)
        ttk.Button(zoom_frame, text="+", width=3,
                  command=self.zoom_in).pack(side='left', padx=2)
        
        # Add search filters with labels
        self.filter1_label = ttk.Label(controls_frame, text="")
        self.filter1_label.pack(pady=5)
        self.processing_filter1_frame = FuzzySearchFrame(controls_frame, width=30, identifier='processing_filter1')
        self.processing_filter1_frame.pack(pady=5)
        
        self.filter2_label = ttk.Label(controls_frame, text="")
        self.filter2_label.pack(pady=5)
        self.processing_filter2_frame = FuzzySearchFrame(controls_frame, width=30, identifier='processing_filter2')
        self.processing_filter2_frame.pack(pady=5)
        
        # Bind filter selection events
        self.processing_filter1_frame.bind('<<ValueSelected>>', lambda e: self.on_filter1_select())
        self.processing_filter2_frame.bind('<<ValueSelected>>', lambda e: None)  # For future use if needed
        
        # Add confirm button
        ttk.Button(controls_frame, text="Confirm", 
                  command=self.process_current_file).pack(pady=20)
        
    def load_excel_data(self):
        try:
            if not all([self.config['excel_file'], self.config['excel_sheet'],
                       self.config['filter1_column'], self.config['filter2_column']]):
                print("Missing configuration values:", {
                    'excel_file': bool(self.config['excel_file']),
                    'excel_sheet': bool(self.config['excel_sheet']),
                    'filter1_column': bool(self.config['filter1_column']),  
                    'filter2_column': bool(self.config['filter2_column'])   
                })
                return
                
            print(f"Loading Excel file: {self.config['excel_file']}")
            print(f"Sheet name: {self.config['excel_sheet']}")
            
            # Load Excel data
            self.excel_data = pd.read_excel(
                self.config['excel_file'],
                sheet_name=self.config['excel_sheet']
            )
            
            print(f"Excel loaded successfully. Columns: {self.excel_data.columns.tolist()}")
            
            # Update filter labels
            self.filter1_label['text'] = self.config['filter1_column']  
            self.filter2_label['text'] = self.config['filter2_column']  
            
            # Get unique values for filters
            self.all_values1 = sorted(self.excel_data[self.config['filter1_column']].unique().tolist())  
            self.all_values2 = sorted(self.excel_data[self.config['filter2_column']].unique().tolist())  
            
            print(f"Found {len(self.all_values1)} unique values for {self.config['filter1_column']}")
            print(f"Found {len(self.all_values2)} unique values for {self.config['filter2_column']}")
            
            # Update filters with values
            self.processing_filter1_frame.set_values(self.all_values1)  
            self.processing_filter2_frame.set_values(self.all_values2)  
            
        except Exception as e:
            self.show_error("Error", f"Error loading Excel data:\n\nAttempted to load:\nFile: {self.config['excel_file']}\nSheet: {self.config['excel_sheet']}\nColumns: {self.config['filter1_column']}, {self.config['filter2_column']}\n\nError: {str(e)}")
            
    def on_filter1_select(self):
        if not self.excel_data is None:
            # Get selected value from first filter
            selected_value = self.processing_filter1_frame.get()
            
            # Filter second filter based on first selection
            filtered_df = self.excel_data[
                self.excel_data[self.config['filter1_column']] == selected_value
            ]
            
            # Update second filter values
            filtered_values = sorted(filtered_df[self.config['filter2_column']].unique().tolist())
            self.processing_filter2_frame.set_values(filtered_values)
            
    def zoom_in(self):
        self.zoom_level = min(3.0, self.zoom_level + 0.2)
        self.display_pdf()
        self.zoom_label.config(text=f"{int(self.zoom_level * 100)}%")
        
    def zoom_out(self):
        self.zoom_level = max(0.2, self.zoom_level - 0.2)
        self.display_pdf()
        self.zoom_label.config(text=f"{int(self.zoom_level * 100)}%")
        
    def process_current_file(self):
        if not self.current_pdf or not os.path.exists(self.current_pdf):
            self.show_error("Error", "No PDF file loaded")
            return
            
        # Create temporary directory for atomic operations
        with tempfile.TemporaryDirectory() as temp_dir:
            try:
                # Get selected values
                value1 = self.processing_filter1_frame.get()
                value2 = self.processing_filter2_frame.get()
                
                if not value1 or not value2:
                    self.show_error("Error", "Please select values from both filters")
                    return
                
                # Find the matching row in Excel
                mask = (self.excel_data[self.config['filter1_column']] == value1) & \
                       (self.excel_data[self.config['filter2_column']] == value2)
                
                if not mask.any():
                    self.show_error("Error", "Selected combination not found in Excel sheet")
                    return
                
                # Get the row data
                row_data = self.excel_data[mask].iloc[0]
                
                # Generate new filename using actual column values
                new_filename = f"{row_data[self.config['filter1_column']]} - {row_data[self.config['filter2_column']]}"
                
                # Clean the filename of invalid characters
                invalid_chars = '<>:"/\\|?*'
                for char in invalid_chars:
                    new_filename = new_filename.replace(char, '_')
                
                # Add .pdf extension if needed
                if not new_filename.lower().endswith('.pdf'):
                    new_filename += '.pdf'
                    
                # Create processed folder if it doesn't exist
                processed_folder = self.config['processed_folder']
                if not os.path.exists(processed_folder):
                    os.makedirs(processed_folder)
                    
                # Generate full path for new file
                new_filepath = os.path.join(processed_folder, new_filename)
                
                # Preview confirmation
                if not messagebox.askyesno("Confirm",
                    f"Current file will be:\n" \
                    f"Renamed to: {new_filename}\n" \
                    f"Moved to: {processed_folder}\n\n" \
                    f"Do you want to proceed?"):
                    return
                    
                # Check if file already exists
                if os.path.exists(new_filepath):
                    if not messagebox.askyesno("Warning", 
                        f"File {new_filename} already exists. Do you want to overwrite it?"):
                        return
                
                # Create backup copies in temp directory
                temp_pdf = os.path.join(temp_dir, "original.pdf")
                temp_excel = os.path.join(temp_dir, "backup.xlsx")
                
                # Copy original files to temp directory
                shutil.copy2(self.current_pdf, temp_pdf)
                shutil.copy2(self.config['excel_file'], temp_excel)
                
                try:
                    # Move and rename the PDF file
                    os.replace(self.current_pdf, new_filepath)
                    
                    # Update Excel with link
                    if self.excel_data is not None:
                        # Get the first matching row index
                        row_idx = mask.idxmax()
                        
                        # Create relative path for Excel link
                        rel_path = os.path.relpath(
                            new_filepath,
                            os.path.dirname(self.config['excel_file'])
                        )
                        
                        # Update Excel file with hyperlink
                        wb = openpyxl.load_workbook(self.config['excel_file'])
                        ws = wb[self.config['excel_sheet']]
                        
                        # Get the last column
                        last_col = ws.max_column
                        
                        # Add header for link column if it doesn't exist
                        if ws.cell(row=1, column=last_col).value != "PDF Link":
                            last_col += 1
                            ws.cell(row=1, column=last_col, value="PDF Link")
                        
                        # Add hyperlink
                        ws.cell(row=row_idx + 2, column=last_col).hyperlink = rel_path
                        
                        # Save Excel file
                        wb.save(self.config['excel_file'])
                        wb.close()
                    
                    messagebox.showinfo("Success", f"File processed successfully: {new_filename}")
                    
                    # Load next PDF only after successful processing
                    self.load_next_pdf()
                    
                except Exception as e:
                    # Restore files from backup
                    if os.path.exists(new_filepath):
                        os.remove(new_filepath)
                    shutil.copy2(temp_pdf, self.current_pdf)
                    shutil.copy2(temp_excel, self.config['excel_file'])
                    
                    raise Exception(f"Error during processing, changes rolled back:\n\nDetails:\n{str(e)}")
                    
            except Exception as e:
                self.show_error("Error", str(e))
            
    def select_source_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.source_folder_entry.delete(0, tk.END)
            self.source_folder_entry.insert(0, folder)
            
    def select_processed_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.processed_folder_entry.delete(0, tk.END)
            self.processed_folder_entry.insert(0, folder)
            
    def select_excel_file(self):
        file = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx;*.xls")])
        if file:
            self.excel_file_entry.delete(0, tk.END)
            self.excel_file_entry.insert(0, file)
            self.update_sheet_list()
            
    def update_sheet_list(self):
        try:
            excel_file = self.excel_file_entry.get()
            if excel_file:
                xl = pd.ExcelFile(excel_file)
                self.sheet_combobox['values'] = xl.sheet_names
                if self.config['excel_sheet'] in xl.sheet_names:
                    self.sheet_combobox.set(self.config['excel_sheet'])
                else:
                    self.sheet_combobox.set(xl.sheet_names[0])
                # Only update column lists if filters exist
                if self.config_filter1_frame and self.config_filter2_frame:
                    self.update_column_lists()
        except Exception as e:
            self.show_error("Error", f"Error reading Excel file:\n\n{str(e)}")
            
    def update_column_lists(self):
        try:
            if not self.config['excel_file'] or not self.sheet_combobox.get():
                return
                
            # Load Excel data
            df = pd.read_excel(
                self.config['excel_file'],
                sheet_name=self.sheet_combobox.get()
            )
            
            # Get all column names
            columns = df.columns.tolist()
            
            # Update filters with column names
            self.config_filter1_frame.set_values(columns)
            if self.config['filter1_column'] in columns:
                self.config_filter1_frame.set(self.config['filter1_column'])
                
            self.config_filter2_frame.set_values(columns)
            if self.config['filter2_column'] in columns:
                self.config_filter2_frame.set(self.config['filter2_column'])
                
        except Exception as e:
            self.show_error("Error", f"Error updating column lists: {str(e)}")
            
    def save_config(self):
        try:
            # Update config with current values
            self.config.update({
                'source_folder': self.source_folder_entry.get(),
                'processed_folder': self.processed_folder_entry.get(),
                'excel_file': self.excel_file_entry.get(),
                'excel_sheet': self.sheet_combobox.get(),
                'filter1_column': self.config_filter1_frame.get(),
                'filter2_column': self.config_filter2_frame.get()
            })
            
            # Ensure config directory exists
            config_dir = os.path.dirname('config.json')
            if config_dir and not os.path.exists(config_dir):
                os.makedirs(config_dir)
            
            # Save with proper formatting
            with open('config.json', 'w') as f:
                json.dump(self.config, f, indent=4)
            messagebox.showinfo("Success", "Configuration saved successfully!")
        except Exception as e:
            self.show_error("Error", f"Error saving configuration:\n\n{str(e)}")
            
    def load_config(self):
        try:
            # Try to load existing config
            if os.path.exists('config.json'):
                with open('config.json', 'r') as f:
                    loaded_config = json.load(f)
                    # Remove old/unused fields
                    if 'filename_pattern' in loaded_config:
                        del loaded_config['filename_pattern']
                    
                    # Update config with loaded values, keeping defaults for missing keys
                    self.config = self.default_config.copy()
                    self.config.update(loaded_config)
            else:
                # Create new config file with defaults if it doesn't exist
                self.save_config()
            
            # Update UI with loaded values
            if hasattr(self, 'source_folder_entry'):
                self.source_folder_entry.delete(0, tk.END)
                self.source_folder_entry.insert(0, self.config['source_folder'])
                
            if hasattr(self, 'processed_folder_entry'):
                self.processed_folder_entry.delete(0, tk.END)
                self.processed_folder_entry.insert(0, self.config['processed_folder'])
                
            if hasattr(self, 'excel_file_entry'):
                self.excel_file_entry.delete(0, tk.END)
                self.excel_file_entry.insert(0, self.config['excel_file'])
                
            if hasattr(self, 'sheet_combobox'):
                if self.config['excel_file']:
                    self.update_sheet_list()
                    
            if hasattr(self, 'config_filter1_frame') and self.config['filter1_column']:
                self.config_filter1_frame.set(self.config['filter1_column'])
                
            if hasattr(self, 'config_filter2_frame') and self.config['filter2_column']:
                self.config_filter2_frame.set(self.config['filter2_column'])
                
        except Exception as e:
            self.show_error("Error", f"Error loading configuration:\n\n{str(e)}")
            # If there's an error, ensure we have default values
            self.config = self.default_config.copy()
            
        # Save cleaned config back to file
        self.save_config()
            
    def load_next_pdf(self):
        try:
            source_folder = self.config['source_folder']
            if not source_folder or not os.path.exists(source_folder):
                self.show_error("Error", f"Source folder not set or does not exist: {source_folder}")
                return
                
            # Get list of PDF files
            pdf_files = [f for f in os.listdir(source_folder) 
                        if f.lower().endswith('.pdf')]
            
            if not pdf_files:
                self.show_error("Error", f"No PDF files found in source folder: {source_folder}")
                return
                
            # If no current PDF, load the first one
            if not self.current_pdf:
                self.current_pdf = os.path.join(source_folder, pdf_files[0])
            else:
                # Find current PDF in the list
                current_name = os.path.basename(self.current_pdf)
                try:
                    current_index = pdf_files.index(current_name)
                    # Load next PDF, or wrap around to first
                    next_index = (current_index + 1) % len(pdf_files)
                    self.current_pdf = os.path.join(source_folder, pdf_files[next_index])
                except ValueError:
                    # Current PDF not found in list, start from beginning
                    self.current_pdf = os.path.join(source_folder, pdf_files[0])
            
            # Display the PDF
            self.display_pdf()
            
            # Update status bar
            current_name = os.path.basename(self.current_pdf)
            total_pdfs = len(pdf_files)
            try:
                current_index = pdf_files.index(current_name) + 1
            except ValueError:
                current_index = 1
            
            self.status_bar.config(
                text=f"PDF {current_index} of {total_pdfs}: {current_name}"
            )
            
        except Exception as e:
            self.show_error("Error", f"Error loading PDF:\n\nSource folder: {source_folder}\nError: {str(e)}")
            self.current_pdf = None
            self.status_bar.config(text="Error loading PDF")
            
    def display_pdf(self):
        try:
            if not self.current_pdf or not os.path.exists(self.current_pdf):
                raise Exception("No PDF file selected or file does not exist")
                
            # Clear previous display
            for widget in self.pdf_frame.winfo_children():
                widget.destroy()
                
            # Open and read the PDF
            doc = fitz.open(self.current_pdf)
            
            if doc.page_count == 0:
                raise Exception("PDF file is empty")
                
            # Get the first page
            page = doc[0]
            
            # Get the pixmap
            mat = fitz.Matrix(2 * self.zoom_level, 2 * self.zoom_level)
            pix = page.get_pixmap(matrix=mat)
            
            # Convert to PhotoImage
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            photo = ImageTk.PhotoImage(img)
            
            # Create and configure canvas
            canvas = tk.Canvas(
                self.pdf_frame,
                width=min(pix.width, 800),
                height=min(pix.height, 600)
            )
            canvas.pack(expand=True, fill='both')
            
            # Add scrollbars
            h_scrollbar = ttk.Scrollbar(
                self.pdf_frame, 
                orient='horizontal',
                command=canvas.xview
            )
            h_scrollbar.pack(side='bottom', fill='x')
            
            v_scrollbar = ttk.Scrollbar(
                self.pdf_frame,
                orient='vertical',
                command=canvas.yview
            )
            v_scrollbar.pack(side='right', fill='y')
            
            # Configure canvas scrolling
            canvas.configure(
                xscrollcommand=h_scrollbar.set,
                yscrollcommand=v_scrollbar.set,
                scrollregion=(0, 0, pix.width, pix.height)
            )
            
            # Display the image
            canvas.create_image(0, 0, anchor='nw', image=photo)
            canvas.image = photo  # Keep a reference
            
            doc.close()
            
        except Exception as e:
            self.show_error("Error", f"Error displaying PDF:\n\n{str(e)}")
            
    def show_error(self, title, message):
        ErrorDialog(self.root, title, message)

def main():
    root = tk.Tk()
    app = FileOrganizerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
