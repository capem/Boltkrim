from __future__ import annotations
from tkinter import StringVar, Widget, Listbox, END, SINGLE, Event, Menu
from tkinter.ttk import Frame, Entry, Style, Scrollbar
from typing import Optional, List, Callable, Any
from difflib import SequenceMatcher
from pathlib import Path
from openpyxl import load_workbook
import os
import subprocess



class FuzzySearchFrame(Frame):
    """Frame containing a fuzzy search entry and listbox."""
    
    def __init__(
        self,
        master: Widget,
        values: Optional[List[str]] = None,
        search_threshold: int = 65,
        identifier: Optional[str] = None,
        on_tab: Optional[Callable[[Event], Optional[str]]] = None,
        **kwargs: Any,
    ) -> None:
        super().__init__(master, **kwargs)

        self.all_values = [str(v) for v in (values or []) if v is not None]
        self.search_threshold = max(
            0, min(100, search_threshold)
        )  # Clamp between 0 and 100
        self.identifier = identifier or "unnamed"
        self.on_tab_callback = on_tab  # Store the tab callback

        # Setup styling
        self._setup_styles()

        self._ignore_next_keyrelease = False

        self._create_widgets()
        self._bind_events()
        self._update_listbox()

        self.after_idle(self.entry.focus_set)

    def _setup_styles(self) -> None:
        """Configure custom styles for the fuzzy search components."""
        style = Style()

        # Entry style
        style.configure(
            "Search.TEntry",
            padding=5,
            relief="solid",
            borderwidth=1,
            font=("Segoe UI", 10),
        )

        # Frame style
        style.configure(
            "Search.TFrame", background="#ffffff", relief="solid", borderwidth=1
        )

    def _create_widgets(self) -> None:
        """Create and configure all child widgets with modern styling."""
        # Entry widget with placeholder
        self.entry_var = StringVar()
        self.entry = Entry(self, style="Search.TEntry", textvariable=self.entry_var)
        self.entry.pack(fill="x", padx=2, pady=2)

        # Set placeholder text
        self._set_placeholder()

        # Listbox frame with scrollbar
        listbox_frame = Frame(self, style="Search.TFrame")
        listbox_frame.pack(fill="both", expand=True, padx=2)

        # Listbox with modern styling
        self.listbox = Listbox(
            listbox_frame,
            height=5,
            exportselection=False,
            selectmode=SINGLE,
            font=("Segoe UI", 10),
            relief="flat",
            background="#ffffff",
            selectbackground="#007bff",
            selectforeground="#ffffff",
            activestyle="none",
        )
        self.listbox.pack(side="left", fill="both", expand=True)

        # Modern scrollbar
        scrollbar = Scrollbar(
            listbox_frame, orient="vertical", command=self.listbox.yview
        )
        scrollbar.pack(side="right", fill="y")
        self.listbox.configure(yscrollcommand=scrollbar.set)

        # Configure mousewheel scrolling
        self._setup_mousewheel_scrolling(listbox_frame, scrollbar)

    def _set_placeholder(self) -> None:
        """Set placeholder text for the entry widget."""
        if not self.entry.get():
            self.entry.configure(foreground="gray")
            self.entry_var.set("Type to search...")

    def _clear_placeholder(self) -> None:
        """Clear placeholder text when entry gets focus."""
        if self.entry.get() == "Type to search...":
            self.entry.configure(foreground="black")
            self.entry_var.set("")

    def _setup_mousewheel_scrolling(self, frame: Frame, scrollbar: Scrollbar) -> None:
        """Setup smooth mousewheel scrolling for the listbox."""

        def _on_mousewheel(event: Event) -> str:
            self.listbox.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"

        def _bind_mousewheel(event: Event) -> None:
            self.listbox.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_mousewheel(event: Event) -> None:
            self.listbox.unbind_all("<MouseWheel>")

        # Bind mousewheel only when mouse is over the listbox area
        self.listbox.bind("<Enter>", _bind_mousewheel)
        self.listbox.bind("<Leave>", _unbind_mousewheel)
        scrollbar.bind("<Enter>", _bind_mousewheel)
        scrollbar.bind("<Leave>", _unbind_mousewheel)

    def _bind_events(self) -> None:
        """Bind widget events to their handlers with improved accessibility."""
        # Entry events
        self.entry.bind("<Key>", self._on_key)  # This will handle all key events
        self.entry.bind("<FocusIn>", self._on_focus_in)
        self.entry.bind("<FocusOut>", self._on_focus_out)
        self.entry.bind("<Return>", lambda e: self._select_top_match() or "break")  # Return "break" to prevent bubbling
        self.entry.bind("<Down>", lambda e: self._focus_listbox())
        self.entry.bind("<Tab>", self._handle_tab)

        # Listbox events
        self.listbox.bind("<<ListboxSelect>>", self._on_select)
        self.listbox.bind("<Button-1>", lambda e: self._on_select)
        self.listbox.bind("<Double-Button-1>", self._on_select)
        self.listbox.bind("<Return>", lambda e: self._on_select(e) or "break")  # Return "break" to prevent bubbling
        self.listbox.bind("<Escape>", lambda e: self.entry.focus_set())
        self.listbox.bind("<Button-3>", self._show_context_menu)  # Right-click binding

    def _on_focus_in(self, event: Optional[Event] = None) -> None:
        """Handle focus-in event with improved visual feedback."""
        self._clear_placeholder()
        self.entry.configure(foreground="black")

    def _on_focus_out(self, event: Optional[Event] = None) -> None:
        """Handle focus-out event with placeholder restoration."""
        if not self.entry.get():
            self._set_placeholder()

    def set_values(self, values: Optional[List[str]]) -> None:
        """Update the list of searchable values."""
        self.all_values = [str(v) for v in (values or []) if v is not None]
        current_value = self.get()  # Use existing get() method which handles placeholder
        self.set(current_value)  # Use existing set() method which handles placeholder
        self._update_listbox()

    def get(self) -> str:
        """Get the current entry text, excluding placeholder."""
        value = self.entry.get()
        return "" if value == "Type to search..." else value

    def set(self, value: str) -> None:
        """Set the entry text with proper placeholder handling."""
        self.entry.delete(0, END)
        if value:
            self.entry.configure(foreground="black")
            self.entry.insert(0, str(value))
        else:
            self._set_placeholder()

    def _on_key(self, event: Event) -> None:
        """Handle all key events in the entry widget."""
        # Skip if this is a modifier key
        if event.keysym in ('Shift_L', 'Shift_R', 'Control_L', 'Control_R', 'Alt_L', 'Alt_R'):
            return
            
        # Schedule an update after the key event is processed
        self.after_idle(self._update_listbox)

    def _update_listbox(self) -> None:
        """Update the listbox with intelligent fuzzy search results."""
        current_value = self.get()  # Use get() to handle placeholder properly

        # Clear current listbox
        self.listbox.delete(0, END)

        # If empty, show all values
        if not current_value:
            for value in self.all_values:
                self.listbox.insert(END, value)
            return

        try:
            search_lower = current_value.lower()
            scored_matches: List[tuple[float, str]] = []

            for value in self.all_values:
                value_lower = value.lower()
                
                # Calculate ratio using SequenceMatcher
                ratio = SequenceMatcher(None, search_lower, value_lower).ratio() * 100
                
                # Apply bonuses for special matches
                if value_lower == search_lower:  # Exact match
                    ratio = 100
                elif value_lower.startswith(search_lower):  # Prefix match
                    ratio = max(ratio, 90)
                elif search_lower in value_lower:  # Contains match
                    ratio = max(ratio, 80)
                elif any(word.startswith(search_lower) for word in value_lower.split()):  # Word boundary match
                    ratio = max(ratio, 75)
                
                # Only include matches that meet the threshold
                if ratio >= self.search_threshold:
                    scored_matches.append((ratio, value))

            # Sort by score (highest first) and add to listbox
            scored_matches.sort(reverse=True, key=lambda x: x[0])
            
            for _, value in scored_matches:
                self.listbox.insert(END, value)

        except Exception as e:
            print(f"Error in fuzzy search ({self.identifier}): {str(e)}")
            # Fall back to simple contains matching
            for value in self.all_values:
                if current_value.lower() in value.lower():
                    self.listbox.insert(END, value)

    def _on_select(self, event: Optional[Event] = None) -> None:
        """Handle selection events in the listbox."""
        if not self.listbox.size():  # If listbox is empty, do nothing
            return

        selection = self.listbox.curselection()
        if selection:
            value = self.listbox.get(selection[0])
            self._select_value(value)
            self.entry.focus_set()

    def _select_value(self, value: str) -> None:
        """Common method to handle value selection and event generation."""
        self.set(value)
        self.event_generate("<<ValueSelected>>")

    def _select_top_match(self) -> None:
        """Select the top match in the listbox when Enter is pressed."""
        if self.listbox.size() > 0:
            self._select_value(self.listbox.get(0))

    def _focus_listbox(self) -> None:
        """Move focus to the listbox when Down arrow is pressed."""
        if self.listbox.size() > 0:
            # Clear any existing selection
            self.listbox.selection_clear(0, END)
            # Set both selection and active item to first item
            self.listbox.selection_set(0)
            self.listbox.activate(0)
            self.listbox.focus_set()
            self.listbox.see(0)  # Ensure the selected item is visible

            # Bind keyboard events when listbox gets focus
            self.listbox.bind("<Up>", self._on_listbox_arrow)
            self.listbox.bind("<Down>", self._on_listbox_arrow)
            self.listbox.bind("<Tab>", self._on_listbox_tab)

    def _on_listbox_tab(self, event: Event) -> Optional[str]:
        """Handle Tab key when pressed in the listbox."""
        active = self.listbox.index("active")
        if active >= 0:
            value = self.listbox.get(active)
            self._select_value(value)
            self.entry.focus_set()
            
            # Call parent's tab handler if provided
            if self.on_tab_callback:
                return self.on_tab_callback(event)
            
            # Default behavior: move to next widget
            event.widget.tk_focusNext().focus()
            return "break"
        return None

    def _on_listbox_arrow(self, event: Event) -> str:
        """Handle up/down arrow keys in listbox to maintain selection."""
        if event.keysym == "Up" and self.listbox.index("active") > 0:
            new_index = self.listbox.index("active") - 1
            self.listbox.selection_clear(0, END)
            self.listbox.selection_set(new_index)
            self.listbox.activate(new_index)
            self.listbox.see(new_index)
        elif (
            event.keysym == "Down"
            and self.listbox.index("active") < self.listbox.size() - 1
        ):
            new_index = self.listbox.index("active") + 1
            self.listbox.selection_clear(0, END)
            self.listbox.selection_set(new_index)
            self.listbox.activate(new_index)
            self.listbox.see(new_index)
        return "break"

    def _handle_tab(self, event: Optional[Event] = None) -> Optional[str]:
        """Handle Tab key press in the entry widget."""
        if self.listbox.winfo_ismapped() and self.listbox.size() > 0:
            # If listbox is visible and has items, select the first one
            self._select_value(self.listbox.get(0))
        
        # Always call parent's tab handler if provided, regardless of listbox state
        if event and self.on_tab_callback:
            return self.on_tab_callback(event)
        
        # Default behavior: move to next widget only if no parent handler
        if event:
            event.widget.tk_focusNext().focus()
        return "break"

    def clear(self) -> None:
        """Clear the entry text and listbox results, resetting to initial state."""
        self.entry.delete(0, END)
        self._set_placeholder()
        self.listbox.delete(0, END)

    def _show_context_menu(self, event: Event) -> None:
        """Show the context menu on right-click."""
        # Get the item at click position
        idx = self.listbox.nearest(event.y)
        if idx < 0 or idx >= self.listbox.size():
            return
        
        # Select the item under cursor
        self.listbox.selection_clear(0, END)
        self.listbox.selection_set(idx)
        self.listbox.activate(idx)
        
        # Get the clicked value
        value = self.listbox.get(idx)
        if not value.startswith("✓ "):  # Only show menu for hyperlinked items
            return
            
        # Create context menu
        menu = Menu(self, tearoff=0)
        menu.add_command(label="Open Linked File", command=lambda: self._open_linked_file(value))
        
        # Show the menu at click position
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _open_linked_file(self, value: str) -> None:
        """Open the linked file from the Excel hyperlink."""
        try:
            # Get ProcessingTab instance
            from .processing_tab import ProcessingTab
            processing_tab = ProcessingTab.get_instance()
            if not processing_tab:
                print("[DEBUG] Could not find ProcessingTab instance")
                return

            _, row_idx = processing_tab.pdf_queue._parse_filter2_value(value)
            if row_idx < 0:
                print(f"[DEBUG] Invalid row index from value: {value}")
                return

            config = processing_tab.config_manager.get_config()
            excel_file = config.get("excel_file", "")
            sheet_name = config.get("excel_sheet", "")
            filter2_col = config.get("filter2_column", "")

            excel_manager = processing_tab.excel_manager
            if not excel_manager._hyperlink_cache.get(row_idx, False):
                print("[DEBUG] No hyperlink found for row")
                return

            # Load workbook to get hyperlink
            wb = load_workbook(excel_file, data_only=True)
            ws = wb[sheet_name]

            # Find the column index
            header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
            if filter2_col not in header:
                print(f"[DEBUG] Column '{filter2_col}' not found")
                return
            col_idx = header[filter2_col]

            # Get cell hyperlink
            cell = ws.cell(row=row_idx + 2, column=col_idx)
            if not cell.hyperlink:
                print("[DEBUG] No hyperlink found in cell")
                return

            # Get full path relative to Excel file location and handle URL encoding
            from urllib.parse import unquote
            excel_dir = Path(excel_file).parent
            decoded_target = unquote(cell.hyperlink.target)
            linked_path = Path(excel_dir) / decoded_target
            
            try:
                normalized_path = linked_path.resolve()
            except Exception:
                # If resolve fails, try with the raw path
                normalized_path = linked_path

            print(f"[DEBUG] Excel dir: {excel_dir}")
            print(f"[DEBUG] Original target: {cell.hyperlink.target}")
            print(f"[DEBUG] Decoded target: {decoded_target}")
            print(f"[DEBUG] Linked path: {linked_path}")
            print(f"[DEBUG] Normalized path: {normalized_path}")

            # Handle both the normalized and raw paths
            paths_to_try = [normalized_path, linked_path]
            file_found = False

            for path_to_check in paths_to_try:
                str_path = str(path_to_check)
                print(f"[DEBUG] Checking path: {str_path}")

                # Check network path availability
                if str_path.startswith("\\\\"):
                    from ..utils.excel_manager import is_path_available
                    if not is_path_available(str_path):
                        print(f"[DEBUG] Network path not accessible: {str_path}")
                        continue

                if path_to_check.exists():
                    file_found = True
                    normalized_path = path_to_check
                    print(f"[DEBUG] Found file at: {normalized_path}")
                    break
                else:
                    print(f"[DEBUG] File not found at: {str_path}")

            if not file_found:
                from .error_dialog import ErrorDialog
                ErrorDialog(self, "Error", f"File not found in any location:\n{str(normalized_path)}\n{str(linked_path)}")
                return

            print(f"[DEBUG] Opening file: {normalized_path}")
            # Open the file using the system default application
            if os.name == 'nt':  # Windows
                os.startfile(normalized_path)
            else:  # Linux/Mac
                subprocess.run(['xdg-open' if os.name == 'posix' else 'open', normalized_path])

        except Exception as e:
            print(f"[DEBUG] Error opening linked file: {str(e)}")
            from .error_dialog import ErrorDialog
            ErrorDialog(self, "Error", f"Error opening linked file: {str(e)}")
        finally:
            if 'wb' in locals():
                wb.close()
