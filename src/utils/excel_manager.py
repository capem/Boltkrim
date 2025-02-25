from os import path, remove
from pandas import read_excel, ExcelFile, DataFrame, Series
from pandas.api.types import is_datetime64_any_dtype
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
from shutil import copy2
from socket import (
    socket,
    AF_INET,
    SOCK_STREAM,
    getdefaulttimeout,
    setdefaulttimeout,
    timeout,
)
from typing import Optional, List, Tuple
from time import sleep
from random import uniform
import pandas as pd
import traceback



def is_path_available(filepath: str, timeout: int = 2) -> bool:
    """Check if a network path is available with timeout.

    Args:
        filepath: The path to check
        timeout: Connection timeout in seconds

    Returns:
        bool: True if path is available, False otherwise
    """
    print(f"[DEBUG] Checking path availability for: {filepath}")

    if not filepath.startswith("\\\\"):  # Not a network path
        exists = path.exists(filepath)
        print(f"[DEBUG] Local path check result: {exists}")
        return exists

    try:
        # Extract server name from UNC path
        server = filepath.split("\\")[2]
        print(f"[DEBUG] Extracted server name: {server}")

        # Try to connect to the server
        print(f"[DEBUG] Attempting connection to {server}:445 with {timeout}s timeout")
        sock = socket(AF_INET, SOCK_STREAM)
        sock.settimeout(timeout)
        sock.connect((server, 445))  # 445 is the SMB port
        sock.close()
        print("[DEBUG] Successfully connected to server")
        return True
    except Exception as e:
        print(f"[DEBUG] Connection failed: {str(e)}")
        return False


def retry_with_backoff(func, max_attempts: int = 5, initial_delay: float = 1.0):
    """Decorator to retry a function with exponential backoff.

    Args:
        func: Function to retry
        max_attempts: Maximum number of retry attempts
        initial_delay: Initial delay between retries in seconds
    """

    def wrapper(*args, **kwargs):
        delay = initial_delay
        last_exception = None

        for attempt in range(max_attempts):
            try:
                return func(*args, **kwargs)
            except (IOError, PermissionError) as e:
                last_exception = e
                if attempt == max_attempts - 1:
                    raise

                # Add jitter to avoid thundering herd
                sleep_time = delay + uniform(0, 0.1 * delay)
                sleep(sleep_time)
                delay *= 2

        raise last_exception

    return wrapper


class ExcelManager:
    """A class to manage Excel file operations with caching and network path handling.

    This class provides functionality to:
    - Load and cache Excel data
    - Update PDF hyperlinks in Excel files
    - Find matching rows based on filter criteria
    - Handle network paths with timeouts and retries
    - Manage Excel sheets and columns

    Attributes:
        excel_data (Optional[DataFrame]): Cached pandas DataFrame containing Excel data
        _cached_file (Optional[str]): Path to the currently cached Excel file
        _cached_sheet (Optional[str]): Name of the currently cached sheet
        _last_modified (Optional[float]): Last modification timestamp of cached file
        _network_timeout (int): Timeout in seconds for network operations
        _hyperlink_cache (Dict[int, bool]): Cache of row indices to hyperlink status
    """

    def __init__(self):
        self.excel_data: Optional[DataFrame] = None
        self._cached_file = None
        self._cached_sheet = None
        self._last_modified = None
        self._network_timeout = 5  # 5 seconds timeout for network operations
        self._hyperlink_cache = {}  # Cache for hyperlink status

    @retry_with_backoff
    def load_excel_data(self, excel_file: str, sheet_name: str) -> bool:
        """Load data from Excel file with caching and retry logic.

        This method loads Excel data while implementing:
        - Network path availability checking
        - Caching to avoid unnecessary reloads
        - Retry logic with exponential backoff
        - Timeout handling for network operations

        Args:
            excel_file: Path to the Excel file to load
            sheet_name: Name of the sheet to load

        Returns:
            bool: True if data was reloaded, False if using cached data

        Raises:
            Exception: If network path is unavailable or Excel file cannot be loaded
        """
        try:
            # Check if we need to reload by comparing file modification time
            try:
                current_modified = path.getmtime(excel_file)
            except (OSError, PermissionError) as e:
                print(f"[DEBUG] Failed to get file modification time: {str(e)}")
                current_modified = None

            # Use cached data if available and not modified
            if (
                self.excel_data is not None
                and self._cached_file == excel_file
                and self._cached_sheet == sheet_name
                and self._last_modified == current_modified
                and current_modified is not None
            ):
                print("[DEBUG] Using cached Excel data")
                return False  # Using cached data

            # Only check path availability if we need to reload
            if not is_path_available(excel_file):
                raise Exception("Network path is not available")

            # Store existing cache before reload
            existing_cache = self._hyperlink_cache.copy() if self._hyperlink_cache else {}
            existing_cache_key = getattr(self, '_last_cached_key', None)
            print(f"[DEBUG] Preserving existing cache - size: {len(existing_cache)}")

            # Load new data with timeout
            original_timeout = getdefaulttimeout()
            setdefaulttimeout(self._network_timeout)
            try:
                print("[DEBUG] Loading fresh Excel data")
                self.excel_data = read_excel(excel_file, sheet_name=sheet_name)
                
                # Restore cache if it was for the same file/sheet
                new_cache_key = f"{excel_file}|{sheet_name}"
                if existing_cache and existing_cache_key and existing_cache_key.startswith(new_cache_key):
                    print("[DEBUG] Restoring preserved cache")
                    self._hyperlink_cache = existing_cache
                    self._last_cached_key = existing_cache_key
                else:
                    print("[DEBUG] Creating new cache")
                    self._hyperlink_cache = {}
                    if hasattr(self, '_last_cached_key'):
                        delattr(self, '_last_cached_key')
            finally:
                setdefaulttimeout(original_timeout)

            # Update cache info
            self._cached_file = excel_file
            self._cached_sheet = sheet_name
            self._last_modified = current_modified

            print(f"[DEBUG] Cache state after reload - size: {len(self._hyperlink_cache)}")
            return True  # Data was reloaded
        except Exception as e:
            if isinstance(e, timeout):
                raise Exception("Network timeout while accessing Excel file")
            raise Exception(f"Error loading Excel data: {str(e)}")

    def cache_hyperlinks_for_column(
        self, excel_file: str, sheet_name: str, column_name: str
    ) -> None:
        """Cache hyperlink information for a specific column."""
        # Skip if Excel data isn't loaded
        if self.excel_data is None:
            print("[DEBUG] Excel data not loaded, skipping hyperlink caching")
            return

        # Skip if hyperlinks are already cached for this file/sheet/column combination
        cache_key = f"{excel_file}|{sheet_name}|{column_name}"
        if hasattr(self, '_last_cached_key') and self._last_cached_key == cache_key and self._hyperlink_cache:
            print(f"[DEBUG] Hyperlinks already cached for {column_name} with {len(self._hyperlink_cache)} entries, skipping")
            return

        print(f"[DEBUG] Loading hyperlink information for column {column_name}")
        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb[sheet_name]
            df_cols = read_excel(excel_file, sheet_name=sheet_name, nrows=0)

            # Find the target column
            if column_name not in df_cols.columns:
                print(f"[DEBUG] Column {column_name} not found in Excel file")
                wb.close()
                return

            target_col = df_cols.columns.get_loc(column_name) + 1
            print(f"[DEBUG] Target column index: {target_col}")

            # Clear existing cache
            old_cache_size = len(self._hyperlink_cache)
            self._hyperlink_cache = {}
            print(f"[DEBUG] Cleared existing cache (size was: {old_cache_size})")

            # Cache hyperlink information only for the target column
            total_rows = len(self.excel_data)
            for idx in range(total_rows):
                cell = ws.cell(
                    row=idx + 2, column=target_col
                )  # +2 for header and 1-based indexing
                has_hyperlink = cell.hyperlink is not None
                self._hyperlink_cache[idx] = has_hyperlink
                # Only print first 10 rows and last row
                if idx < 10 or idx == total_rows - 1:
                    print(f"[DEBUG] Row {idx}: Hyperlink status = {has_hyperlink}")
                elif idx == 10:
                    print(f"[DEBUG] ... ({total_rows - 11} more rows) ...")

            # Store the cache key
            self._last_cached_key = cache_key
            print(f"[DEBUG] Updated cache key to: {cache_key}")
            print(f"[DEBUG] New cache size: {len(self._hyperlink_cache)}")

            wb.close()
            print(f"[DEBUG] Successfully cached hyperlink information for column {column_name}")
        except Exception as e:
            print(f"[DEBUG] Error caching hyperlink information: {str(e)}")
            print(f"[DEBUG] Stack trace: {traceback.format_exc()}")
            # Don't clear the cache if there was an error
            if not self._hyperlink_cache and old_cache_size > 0:
                print("[DEBUG] Restoring previous cache after error")
                self._hyperlink_cache = {}

    @retry_with_backoff
    def update_pdf_link(
        self,
        excel_file: str,
        sheet_name: str,
        row_idx: int,
        pdf_path: str,
        filter2_col: str,
    ) -> Optional[str]:
        """Update Excel with PDF link and capture the original hyperlink.

        Args:
            excel_file: Path to the Excel file
            sheet_name: Name of the sheet to update
            row_idx: Row index to update (0-based)
            pdf_path: Path to the PDF file
            filter2_col: Column name for the hyperlink

        Returns:
            Optional[str]: The original hyperlink if it existed, else None
        """
        try:
            print(f"[DEBUG] Cache state before PDF link update - size: {len(self._hyperlink_cache)}")
            if not is_path_available(excel_file):
                raise FileNotFoundError(f"Excel file not found or not accessible: {excel_file}")

            if not is_path_available(pdf_path):
                raise FileNotFoundError(f"PDF file not found or not accessible: {pdf_path}")

            print(f"[DEBUG] Updating Excel link in {sheet_name}, row {row_idx + 2}, column {filter2_col}")

            wb = None
            backup_created = False
            original_hyperlink = None

            try:
                wb = load_workbook(excel_file, data_only=False)
                ws = wb[sheet_name]

                header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
                if filter2_col not in header:
                    raise ValueError(f"Column '{filter2_col}' not found in sheet '{sheet_name}'")
                col_idx = header[filter2_col]

                cell = ws.cell(row=row_idx + 2, column=col_idx)

                if cell.hyperlink:
                    original_hyperlink = cell.hyperlink.target
                    print(f"[DEBUG] Found existing hyperlink: {original_hyperlink}")

                if not backup_created:
                    backup_file = f"{excel_file}.bak"
                    copy2(excel_file, backup_file)
                    backup_created = True
                    print(f"[DEBUG] Backup created at {backup_file}")

                relative_pdf_path = path.relpath(pdf_path, path.dirname(excel_file))
                print(f"[DEBUG] Setting new hyperlink: {relative_pdf_path}")

                hyperlink = Hyperlink(
                    ref=cell.coordinate,
                    target=relative_pdf_path,
                    display=cell.value,
                )
                cell.hyperlink = hyperlink
                cell.style = 'Hyperlink'

                print("[DEBUG] Saving workbook")
                wb.save(excel_file)
                
                # Update the hyperlink cache for this row
                self._hyperlink_cache[row_idx] = True
                print(f"[DEBUG] Updated hyperlink cache for row {row_idx} to True")
                print(f"[DEBUG] Cache state after PDF link update - size: {len(self._hyperlink_cache)}")

                print(f"[DEBUG] Excel file updated with new hyperlink at row {row_idx + 2}, column {filter2_col}")

                return original_hyperlink

            except Exception as e:
                print(f"[DEBUG] Error in update_pdf_link: {str(e)}")
                if wb:
                    try:
                        print("[DEBUG] Closing workbook without saving due to error")
                        wb.close()
                    except IOError:  # For workbook close operations
                        pass

                if backup_created and path.exists(backup_file):
                    try:
                        print("[DEBUG] Restoring from backup")
                        copy2(backup_file, excel_file)
                    except (IOError, OSError):  # For file copy operations
                        pass
                raise Exception(f"Error updating Excel with PDF link: {str(e)}")

            finally:
                if wb:
                    wb.close()

        except Exception as e:
            print(f"[DEBUG] Error in update_pdf_link: {str(e)}")
            print(f"[DEBUG] Stack trace: {traceback.format_exc()}")
            raise Exception(f"Error updating Excel with PDF link: {str(e)}")

    @retry_with_backoff
    def revert_pdf_link(
        self,
        excel_file: str,
        sheet_name: str,
        row_idx: int,
        filter2_col: str,
        original_hyperlink: str,
        original_value: str,
    ) -> None:
        """Revert the Excel cell's hyperlink and value to its original state."""
        try:
            print(f"[DEBUG] Attempting to revert Excel cell at row {row_idx}, column {filter2_col}")
            
            # Load workbook with data_only=False to preserve all formulas in the workbook
            wb = load_workbook(excel_file, data_only=False)
            ws = wb[sheet_name]
            
            # Map headers to column indices (1-based)
            header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
            if filter2_col not in header:
                raise ValueError(f"Column '{filter2_col}' not found in sheet '{sheet_name}'")
            col_idx = header[filter2_col]
            
            # Get the cell to update
            cell = ws.cell(row=row_idx + 2, column=col_idx)  # +2 for header and 1-based indexing
            
            # Clear existing hyperlink
            if cell.hyperlink:
                cell.hyperlink = None
            
            # Reset the cell value
            cell.value = original_value
            
            if original_hyperlink:
                # If there was an original hyperlink, restore it and keep hyperlink style
                cell.hyperlink = Hyperlink(
                    ref=cell.coordinate,
                    target=original_hyperlink
                )
                cell.font = Font(underline="single", color="0000FF")
                print(f"[DEBUG] Restored hyperlink and style: {original_hyperlink}")
            else:
                # If there was no original hyperlink, clear the hyperlink style
                cell.font = Font()
                print("[DEBUG] Cleared hyperlink style")
            
            # Save the workbook
            wb.save(excel_file)
            print("[DEBUG] Successfully reverted Excel cell")
            
        except Exception as e:
            print(f"[DEBUG] Error reverting Excel cell: {str(e)}")
            raise Exception(f"Failed to revert Excel cell: {str(e)}")
        finally:
            if wb:
                wb.close()

    def get_sheet_names(self, excel_file: str) -> List[str]:
        """Get list of sheet names from Excel file."""
        try:
            if not is_path_available(excel_file):
                raise Exception("Network path is not available")

            original_timeout = getdefaulttimeout()
            setdefaulttimeout(self._network_timeout)
            try:
                xl = ExcelFile(excel_file)
                return xl.sheet_names
            finally:
                setdefaulttimeout(original_timeout)
        except Exception as e:
            if isinstance(e, timeout):
                raise Exception("Network timeout while accessing Excel file")
            raise Exception(f"Error reading Excel sheets: {str(e)}")

    def get_column_names(self) -> List[str]:
        """Get list of column names from loaded Excel data."""
        if self.excel_data is None:
            return []
        return list(self.excel_data.columns)

    def find_matching_row(
        self,
        filter_columns: List[str],
        filter_values: List[str],
    ) -> Tuple[Optional[Series], Optional[int]]:
        """Find a row matching the given filter values.

        Args:
            filter_columns: List of column names to filter on
            filter_values: List of values to match in corresponding columns

        Returns:
            Tuple[Optional[Series], Optional[int]]: Matching row data and index if found
        """
        try:
            if self.excel_data is None:
                raise Exception("Excel data not loaded")

            if len(filter_columns) != len(filter_values):
                raise Exception(f"Number of filter columns ({len(filter_columns)}) and values ({len(filter_values)}) must match")

            print(f"[DEBUG] Finding row with filters: {dict(zip(filter_columns, filter_values))}")

            df = self.excel_data

            def create_mask(col: str, value: str) -> Series:
                """Create a boolean mask for filtering DataFrame."""
                if col not in df.columns:
                    raise Exception(f"Column '{col}' not found in Excel file. Available columns: {', '.join(df.columns)}")

                print(f"[DEBUG] Creating mask for column '{col}' with value '{value}'")

                # Make a copy of the column to avoid modifying the original
                col_series = df[col].copy()

                # Handle date columns
                if "DATE" in col.upper():
                    try:
                        # Try to parse the date value with dayfirst=True for dd/mm/yyyy format
                        print(f"[DEBUG] Attempting to parse date value '{value}' for column '{col}'")
                        # First try explicit French format
                        try:
                            date_value = pd.to_datetime(value, format='%d/%m/%Y')
                        except ValueError:  # For date parsing errors
                            # Fallback to general parsing with dayfirst=True
                            date_value = pd.to_datetime(value, dayfirst=True)
                        
                        print(f"[DEBUG] Successfully parsed input date: {date_value}")
                        
                        # Convert column to datetime if not already
                        if not is_datetime64_any_dtype(col_series):
                            # Try to parse column values with French format first
                            try:
                                col_series = pd.to_datetime(col_series, format='%d/%m/%Y', errors='coerce')
                            except ValueError:  # For date parsing errors
                                col_series = pd.to_datetime(col_series, dayfirst=True, errors='coerce')
                            
                        print(f"[DEBUG] Column conversion complete. Sample values: {col_series.head()}")
                        return col_series.dt.normalize() == date_value.normalize()
                        
                    except Exception as e:
                        print(f"[DEBUG] Date parsing failed for column '{col}': {str(e)}")
                        print("[DEBUG] Falling back to string comparison")
                        return col_series.astype(str).str.strip() == str(value).strip()
                
                # Handle numeric columns (for amount comparisons)
                elif "MNT" in col.upper() or any(num_indicator in col.upper() for num_indicator in ["MONTANT", "NOMBRE", "NUM", "PRIX"]):
                    try:
                        # Clean and convert the input value
                        num_str = str(value).replace(' ', '').replace(',', '.')
                        target_value = float(num_str)
                        
                        # Convert column values with the same cleaning
                        def clean_number(x):
                            try:
                                if pd.isna(x):
                                    return None
                                return float(str(x).replace(' ', '').replace(',', '.'))
                            except (ValueError, TypeError):  # For number conversion errors
                                return None
                        
                        col_series = col_series.apply(clean_number)
                        return col_series == target_value
                        
                    except Exception as e:
                        print(f"[DEBUG] Number parsing failed for column '{col}': {str(e)}")
                        return col_series.astype(str).str.strip() == str(value).strip()
                
                # Default string comparison
                else:
                    return col_series.astype(str).str.strip() == str(value).strip()

            # Create mask for each filter
            masks = [create_mask(col, val) for col, val in zip(filter_columns, filter_values)]
            
            # Combine all masks with AND operation
            final_mask = pd.Series([True] * len(df))
            for mask in masks:
                final_mask &= mask

            matching_rows = df[final_mask]
            print(f"[DEBUG] Found {len(matching_rows)} matching rows")

            if len(matching_rows) == 0:
                print("[DEBUG] No matching rows found")
                return None, None
            elif len(matching_rows) == 1:
                row_idx = matching_rows.index[0]
                print(f"[DEBUG] Found exactly one match at row index {row_idx}")
                return matching_rows.iloc[0], row_idx
            else:
                # If multiple matches, return the first one
                row_idx = matching_rows.index[0]
                print(f"[DEBUG] Found multiple matches, using first match at row index {row_idx}")
                return matching_rows.iloc[0], row_idx

        except Exception as e:
            print(f"[DEBUG] Error in find_matching_row: {str(e)}")
            raise Exception(f"Error finding matching row: {str(e)}")

    def has_hyperlink(self, row_idx: int) -> bool:
        """Check if a row has any hyperlinks.

        Args:
            row_idx: The 0-based row index to check

        Returns:
            bool: True if the row has any hyperlinks, False otherwise
        """
        return self._hyperlink_cache.get(row_idx, False)

    @retry_with_backoff
    def add_new_row(
        self,
        excel_file: str,
        sheet_name: str,
        filter_columns: List[str],
        filter_values: List[str],
    ) -> Tuple[Series, int]:
        try:
            print(f"[DEBUG] Cache state before adding new row - size: {len(self._hyperlink_cache)}")
            if len(filter_columns) != len(filter_values):
                raise Exception(f"Number of filter columns ({len(filter_columns)}) and values ({len(filter_values)}) must match")

            print(f"[DEBUG] Adding new row with values: {dict(zip(filter_columns, filter_values))}")

            # Create a temporary backup
            backup_file = excel_file + ".bak"
            copy2(excel_file, backup_file)
            print(f"[DEBUG] Created backup at {backup_file}")

            try:
                # Load workbook
                wb = load_workbook(excel_file)
                ws = wb[sheet_name]

                # Get header row to map column names to indices
                header_row = ws[1]
                col_indices = {cell.value: idx + 1 for idx, cell in enumerate(header_row)}

                # Verify all filter columns exist
                for col in filter_columns:
                    if col not in col_indices:
                        raise Exception(f"Column '{col}' not found in Excel file. Available columns: {', '.join(col_indices.keys())}")

                # Find the first table's range to determine where to add the new row
                table_end_row = None
                for table in ws.tables.values():
                    try:
                        current_ref = table.ref
                        ref_parts = current_ref.split(':')
                        if len(ref_parts) == 2:
                            end_ref = ref_parts[1]
                            table_end_row = int(''.join(filter(str.isdigit, end_ref)))
                            print(f"[DEBUG] Found table with end row: {table_end_row}")
                            break  # Use the first table found
                    except Exception as e:
                        print(f"[DEBUG] Error processing table reference: {str(e)}")
                        continue

                # If we found a table, add the row immediately after it
                if table_end_row:
                    new_row_idx = table_end_row + 1
                else:
                    # Fallback to adding at the end if no table found
                    new_row_idx = ws.max_row + 1
                
                print(f"[DEBUG] Adding row at index {new_row_idx - 2} (Excel row {new_row_idx})")

                # First pass: Copy all formats from the template row (use row before new row)
                template_row = new_row_idx - 1
                print(f"[DEBUG] Using template row {template_row} for formatting")
                for col_idx in range(1, len(header_row) + 1):
                    template_cell = ws.cell(row=template_row, column=col_idx)
                    new_cell = ws.cell(row=new_row_idx, column=col_idx)
                    
                    # Copy number format and style
                    new_cell.number_format = template_cell.number_format
                    new_cell._style = template_cell._style

                # Second pass: Set values with proper type conversion
                for col, val in zip(filter_columns, filter_values):
                    col_idx = col_indices[col]
                    template_cell = ws.cell(row=template_row, column=col_idx)
                    new_cell = ws.cell(row=new_row_idx, column=col_idx)
                    
                    # Convert value based on the column type
                    if "DATE" in col.upper() and val:
                        try:
                            # Try to parse date in French format first (dd/mm/yyyy)
                            french_date_formats = ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%d_%m_%Y']
                            date_val = None
                            
                            for fmt in french_date_formats:
                                try:
                                    date_val = pd.to_datetime(val, format=fmt)
                                    break
                                except ValueError:  # For date parsing errors
                                    continue
                                    
                            if date_val is None:
                                # Fallback to pandas default parser
                                date_val = pd.to_datetime(val)
                                
                            new_cell.value = date_val.to_pydatetime()
                            # Set French date format
                            new_cell.number_format = 'DD/MM/YYYY'
                        except (ValueError, TypeError):  # For date parsing errors
                            new_cell.value = val
                            print(f"[DEBUG] Could not parse date '{val}' for column '{col}'")
                            
                    elif "MNT" in col.upper() or any(num_indicator in col.upper() for num_indicator in ["MONTANT", "NOMBRE", "NUM", "PRIX"]):
                        try:
                            # Handle French number format (comma as decimal separator)
                            if isinstance(val, str):
                                # Replace comma with dot for conversion, handle thousands separator
                                num_str = val.replace(' ', '').replace(',', '.')
                                num_val = float(num_str)
                                new_cell.value = num_val
                            else:
                                new_cell.value = float(val)
                        except (ValueError, TypeError):  # For number conversion errors
                            new_cell.value = val
                            print(f"[DEBUG] Could not parse number '{val}' for column '{col}'")
                    else:
                        new_cell.value = val
                    
                    print(f"[DEBUG] Set value '{val}' for column '{col}'")
                    
                # Save workbook
                # Check and expand table ranges to include the new row
                print("[DEBUG] Checking for tables that need to be expanded")
                for table in ws.tables.values():
                    try:
                        current_ref = table.ref
                        # Split table reference into components (e.g., 'A1:D10' -> ['A1', 'D10'])
                        ref_parts = current_ref.split(':')
                        if len(ref_parts) != 2:
                            continue
                            
                        start_ref, end_ref = ref_parts
                        # Extract row numbers from references
                        start_row = int(''.join(filter(str.isdigit, start_ref)))
                        end_row = int(''.join(filter(str.isdigit, end_ref)))
                        
                        # Check if new row is immediately after table
                        if end_row == new_row_idx - 1:
                            # Get column letters from references (e.g., 'A' from 'A1')
                            start_col = ''.join(filter(str.isalpha, start_ref))
                            end_col = ''.join(filter(str.isalpha, end_ref))
                            
                            # Create new reference that includes the new row
                            new_ref = f"{start_col}{start_row}:{end_col}{new_row_idx}"
                            table.ref = new_ref
                            print(f"[DEBUG] Expanded table '{table.displayName}' range to {new_ref}")
                    except Exception as table_e:
                        print(f"[DEBUG] Error expanding table: {str(table_e)}")
                        # Continue with other tables even if one fails
                        continue

                # Save workbook with updated table ranges
                wb.save(excel_file)
                
                # Update cache for the new row
                self._hyperlink_cache[new_row_idx - 2] = False
                print(f"[DEBUG] Updated hyperlink cache for new row {new_row_idx - 2}")
                print(f"[DEBUG] Cache state after adding row - size: {len(self._hyperlink_cache)}")

                # Create new row data for return
                new_row_data = pd.Series(index=self.excel_data.columns, dtype='object')
                for col, val in zip(filter_columns, filter_values):
                    new_row_data[col] = val
                self.excel_data = pd.concat([self.excel_data, pd.DataFrame([new_row_data]).dropna(how='all', axis=1)], ignore_index=True)

                print(f"[DEBUG] Successfully added new row at index {new_row_idx - 2}")
                
                # Remove backup after successful write
                if path.exists(backup_file):
                    remove(backup_file)
                    print("[DEBUG] Removed backup file after successful write")

                return new_row_data, new_row_idx - 2

            finally:
                if wb:
                    wb.close()

        except Exception as e:
            print(f"[DEBUG] Error in add_new_row: {str(e)}")
            print(f"[DEBUG] Stack trace: {traceback.format_exc()}")
            if path.exists(backup_file):
                try:
                    copy2(backup_file, excel_file)
                    print("[DEBUG] Restored from backup after error")
                except (IOError, OSError):  # For file copy operations
                    print("[DEBUG] Failed to restore from backup")
            raise Exception(f"Error adding new row: {str(e)}")
